from typing import Any
import os
from openpyxl import load_workbook
import threading
from utils import logger
import time

WORK_DIRECTORY = os.path.dirname(os.path.abspath(__file__))

lock = threading.Lock()

MAX_RETRIES = 30
RETRY_DELAY = 2  # секунды

def get_col_by_name(filename: str, column_name: str):
    file_path = os.path.join(WORK_DIRECTORY, filename)
    for attempt in range(MAX_RETRIES):
        if True:
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active

                for i, cell in enumerate(ws[1], start=1):
                    if cell.value == column_name:
                        wb.close()
                        return i  # номер колонки

                wb.close()
                return None  # колонка не найдена

            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
                else:
                    raise Exception(f"Не удалось получить колонку '{column_name}' из файла после {MAX_RETRIES} попыток: {e}")

def add_column_if_not_exists(filename: str, column_name: str):
    file_path = os.path.join(WORK_DIRECTORY, filename)

    for attempt in range(MAX_RETRIES):
        if True:
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active

                existing_columns = [cell.value for cell in ws[1] if cell.value]

                if column_name not in existing_columns:
                    new_column_number = ws.max_column + 1
                    ws.cell(row=1, column=new_column_number, value=column_name)
                    wb.save(file_path)

                wb.close()
                return  # успешно

            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
                else:
                    raise Exception(f"Не удалось добавить колонку '{column_name}' после {MAX_RETRIES} попыток: {e}")

def write_cell(filename: str, column_name: str, profile_number: Any, write_value: Any):
    file_path = os.path.join(WORK_DIRECTORY, filename)

    for attempt in range(MAX_RETRIES):
        with lock:
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active

                col_num = -1
                try:
                    col_num = get_col_by_name(filename, column_name)
                except:
                    pass

                if col_num == -1:
                    add_column_if_not_exists(filename, column_name)

                row_num = 0
                for index, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
                    if row[0].value == profile_number:
                        row_num = index
                        break

                if row_num > 0 and col_num:
                    ws.cell(row=row_num, column=col_num, value=write_value)

                wb.save(file_path)
                wb.close()
                return  # успешно

            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    time.sleep(RETRY_DELAY)
                else:
                    raise Exception(f"Не удалось записать значение в ячейку: {e}")

def is_numeric(value):
    try:
        int(value)
        return True
    except (TypeError, ValueError):
        return False

def get_profile_for_work(filename: str):
    file_path = os.path.join(WORK_DIRECTORY, filename)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel файл не найден рядом со скриптом!")

    for attempt in range(MAX_RETRIES):
        try:
            with lock:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active

                headers = [cell.value for cell in ws[1]]
                header_index = {name: idx for idx, name in enumerate(headers)}

                required_cols = ["NUMBER_WALLET", "STATUS"]
                if not all(col in header_index for col in required_cols):
                    raise ValueError("Не найдены нужные колонки (NUMBER_WALLET / STATUS)")

                filtered_data = []

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}

                    number_profile = row_dict.get("NUMBER_WALLET")
                    profile_status = row_dict.get("STATUS")

                    if (profile_status is None or str(profile_status).strip() == "") and is_numeric(number_profile):
                        filtered_data.append(row_dict)

                wb.close()
                return filtered_data

        except Exception as e:
            if attempt == MAX_RETRIES - 1:
                logger.error(f"Ошибка при работе с файлом (попытка {attempt + 1}): {e}")
            time.sleep(RETRY_DELAY)

    raise Exception("Не удалось обработать Excel после нескольких попыток")

